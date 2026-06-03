"""
Escreve a planilha de saída como .xlsx.
Regra R1: SOMENTE a coluna 'Descrição complementar' é alterada.
Regra R5: Backup automático antes de salvar.
"""

import shutil
from datetime import datetime
from pathlib import Path

import openpyxl


def salvar_planilha(
    dados_originais: dict,
    descricoes: dict,          # {sku: html_string}
    output_dir: str,
    backup_dir: str,
) -> tuple[str, str, int]:
    """
    Grava a planilha final e faz backup do original.

    Args:
        dados_originais: retorno de leitor.ler_planilha()
        descricoes:      {sku: html} dos pais processados com sucesso
        output_dir:      pasta de saída
        backup_dir:      pasta de backup

    Returns:
        (caminho_output, caminho_backup, qtd_celulas_alteradas)
    """
    original_path = Path(dados_originais["filepath"])

    # R5 — backup antes de qualquer coisa
    backup_path = _fazer_backup(original_path, backup_dir)

    # Monta o novo xlsx
    output_path = _escrever_xlsx(dados_originais, descricoes, output_dir, original_path)

    # Verificação de integridade (R1)
    alteracoes = _verificar_integridade(dados_originais, descricoes, output_path)

    return str(output_path), str(backup_path), alteracoes


# ---------------------------------------------------------------------------
# Internos
# ---------------------------------------------------------------------------

def _fazer_backup(original: Path, backup_dir: str) -> Path:
    pasta = Path(backup_dir)
    pasta.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = pasta / f"{original.stem}_backup_{ts}{original.suffix}"
    shutil.copy2(original, destino)
    return destino


def _escrever_xlsx(dados: dict, descricoes: dict, output_dir: str, original: Path) -> Path:
    pasta = Path(output_dir)
    pasta.mkdir(parents=True, exist_ok=True)

    nome_saida = f"{original.stem}_FINAL.xlsx"
    destino = pasta / nome_saida

    headers  = dados["headers"]
    rows     = dados["rows"]
    ci       = dados["col_index"]
    col_sku  = ci["Código (SKU)"]
    col_desc = ci["Descrição complementar"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produtos"

    # Cabeçalho
    ws.append(headers)

    # Dados
    for row in rows:
        sku = str(row[col_sku]).strip()
        nova_linha = list(row)
        if sku in descricoes:
            nova_linha[col_desc] = descricoes[sku]
        ws.append(nova_linha)

    wb.save(destino)
    return destino


def _verificar_integridade(dados: dict, descricoes: dict, output_path: Path) -> int:
    """
    Confere que NENHUMA coluna além de 'Descrição complementar' foi alterada.
    Lança RuntimeError se encontrar divergência.
    Retorna o número de células corretamente alteradas.
    """
    headers  = dados["headers"]
    rows_in  = dados["rows"]
    ci       = dados["col_index"]
    col_sku  = ci["Código (SKU)"]
    col_desc = ci["Descrição complementar"]

    wb = openpyxl.load_workbook(str(output_path), read_only=True)
    ws = wb.active
    rows_out = list(ws.iter_rows(values_only=True))
    wb.close()

    # Linha 0 da planilha é o cabeçalho; dados começam na linha 1
    alteracoes = 0
    for i, row_orig in enumerate(rows_in):
        row_saida = rows_out[i + 1]  # +1 pelo cabeçalho
        sku = str(row_orig[col_sku]).strip()

        for j in range(len(row_orig)):
            val_orig  = str(row_orig[j]).strip()
            val_saida = str(row_saida[j]).strip() if row_saida[j] is not None else ""

            if j == col_desc:
                if sku in descricoes:
                    alteracoes += 1
                # Não verifica a coluna de destino — ela é esperada divergir
                continue

            if val_orig != val_saida:
                raise RuntimeError(
                    f"INTEGRIDADE: coluna '{headers[j]}' (índice {j}) foi "
                    f"modificada na linha {i + 2} (SKU: {sku!r}).\n"
                    f"  Original: {val_orig!r}\n"
                    f"  Saída:    {val_saida!r}"
                )

    return alteracoes
