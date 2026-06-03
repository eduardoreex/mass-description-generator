"""
Lê planilhas .xls e .xlsx exportadas do Tiny ERP.
Retorna um dict com headers, rows e col_index.
"""

import warnings
from pathlib import Path

# Colunas obrigatórias para o sistema funcionar
COLUNAS_OBRIGATORIAS = [
    "Código (SKU)",
    "Descrição",
    "URL imagem 1",
    "Descrição complementar",
]


def ler_planilha(filepath: str) -> dict:
    """
    Abre .xls ou .xlsx e retorna:
    {
        'headers': [...],
        'rows': [[...], ...],
        'col_index': {'NomeColuna': indice, ...},
        'filepath': str,
        'total_linhas': int,
    }
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")

    ext = path.suffix.lower()
    if ext == ".xls":
        dados = _ler_xls(path)
    elif ext == ".xlsx":
        dados = _ler_xlsx(path)
    else:
        raise ValueError(f"Formato não suportado: '{ext}'. Use .xls ou .xlsx")

    _validar_colunas(dados["col_index"], filepath)
    return dados


# ---------------------------------------------------------------------------
# Leitores internos
# ---------------------------------------------------------------------------

def _ler_xls(path: Path) -> dict:
    try:
        import xlrd
    except ImportError:
        raise ImportError("Instale o xlrd: pip install xlrd>=2.0.1")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = xlrd.open_workbook(str(path), ignore_workbook_corruption=True)

    ws = wb.sheet_by_index(0)

    if ws.nrows < 2:
        raise ValueError("Planilha vazia ou com apenas cabeçalho — sem linhas de dados.")

    headers = [_cel_para_str_xls(ws.cell(0, c)) for c in range(ws.ncols)]

    rows = []
    for r in range(1, ws.nrows):
        row = [_cel_para_str_xls(ws.cell(r, c)) for c in range(ws.ncols)]
        rows.append(row)

    return _montar_resultado(headers, rows, str(path))


def _ler_xlsx(path: Path) -> dict:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Instale o openpyxl: pip install openpyxl>=3.1.2")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    todas = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(todas) < 2:
        raise ValueError("Planilha vazia ou com apenas cabeçalho — sem linhas de dados.")

    headers = [str(v).strip() if v is not None else "" for v in todas[0]]

    rows = []
    for linha in todas[1:]:
        row = [str(v).strip() if v is not None else "" for v in linha]
        # Garante que todas as linhas têm o mesmo número de colunas
        while len(row) < len(headers):
            row.append("")
        rows.append(row)

    return _montar_resultado(headers, rows, str(path))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cel_para_str_xls(cell) -> str:
    import xlrd
    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return ""
    if cell.ctype == xlrd.XL_CELL_TEXT:
        return str(cell.value).strip()
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        v = cell.value
        return str(int(v)) if v == int(v) else str(v)
    return str(cell.value).strip()


def _montar_resultado(headers: list, rows: list, filepath: str) -> dict:
    col_index = {}
    for i, h in enumerate(headers):
        h_limpo = h.strip()
        if h_limpo and h_limpo not in col_index:  # primeira ocorrência vence
            col_index[h_limpo] = i

    return {
        "headers":      headers,
        "rows":         rows,
        "col_index":    col_index,
        "filepath":     filepath,
        "total_linhas": len(rows),
    }


def _validar_colunas(col_index: dict, filepath: str) -> None:
    faltando = [c for c in COLUNAS_OBRIGATORIAS if c not in col_index]
    if faltando:
        disponiveis = list(col_index.keys())[:15]
        raise ValueError(
            f"Colunas obrigatórias não encontradas em '{filepath}':\n"
            f"  Faltando: {faltando}\n"
            f"  Disponíveis (primeiras 15): {disponiveis}"
        )
